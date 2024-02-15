from itertools import chain

from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.core.exceptions import FieldError

from collections import defaultdict
from main import models
from main.models import Product 
from . funcs import search_with_fields, pagenator_page

import xlwt
import openpyxl
from datetime import datetime
from sqlalchemy import create_engine, Table, Column, Integer, String, DateTime
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker


def dashboard(request):
    categorys = models.Category.objects.all()
    products = models.Product.objects.all()
    users = User.objects.all()
    context = {
        'categorys':categorys,
        'products':products,
        'users':users,
    }
    return render(request, 'dashboard/index.html', context)


def category_list(request):
    categorys = models.Category.objects.all()
    return render(request, 'category/list.html', {'categorys':categorys})


def category_detail(request, id):
    category = models.Category.objects.get(id=id)
    products = models.Product.objects.filter(category=category, is_active=True)
    context = {
        'category':category,
        'products':products
    }
    return render(request, 'category/list.html', context)


def category_update(request, id):
    category = models.Category.objects.get(id=id)
    category.name = request.POST['name']
    category.save()
    return redirect('category_detail', category.id)


def category_delete(request, id):
    category = models.Category.objects.get(id=id)
    category.delete()
    return redirect('category_list')

# products

def products(request):
    products = models.Product.objects.all()
    return render(request, 'dashboard/products/list.html', {'products':products})


def product_create(request):
    categorys = models.Category.objects.all()
    context = {
        'categorys':categorys
    }
    if request.method == "POST":
        name = request.POST['name']
        description = request.POST['description']
        quantity = request.POST['quantity']
        price = request.POST['price']
        currency = request.POST['currency']
        baner_image = request.FILES['baner_image']
        category_id = request.POST['category_id']
        images = request.FILES.getlist('images')
        product = models.Product.objects.create(
            name=name,
            description = description,
            quantity=quantity,
            price=price,
            currency=currency,
            baner_image=baner_image,
            category_id=category_id
        )
        for image in images:
            models.ProductImage.objects.create(
                image=image,
                product=product
            )

    return render(request, 'dashboard/products/create.html', context)


def create_enter(request):
    if request.method == 'POST':
        product_id = request.POST['product_id']
        quantity = int(request.POST['quantity'])
        models.EnterProduct.objects.create(
            product_id=product_id,
            quantity=quantity
        )
        return redirect('dashboard:list_enter')
    return render(request, 'dashboard/enter/create.html', {'products':models.Product.objects.all()})


def update_enter(request, id):
    if request.method == 'POST':
        quantity = int(request.POST['quantity'])
        enter = models.EnterProduct.objects.get(id=id)
        enter.quantity = quantity
        enter.save()
    return redirect('dashboard:list_enter')


def delete_enter(request, id):
    models.EnterProduct.objects.get(id=id).delete()
    return redirect('dashboard:list_enter')


def list_enter(request):
    enters = models.EnterProduct.objects.all()
    context = {'enters':enters}
    return render(request, 'dashboard/enter/list.html', context)

#chiqim

def expenditure(request):
    cartproducts = models.CartProduct.objects.filter(card__is_active=False)
    
    dict1 = defaultdict(int)
    for cartproduct in cartproducts:
        dict1[cartproduct.product.name] += cartproduct.quantity
    result_list = [{'name': name, 'total_quantity': total_quantity} for name, total_quantity in dict1.items()]

    return render(request, 'dashboard/chiqim/list.html', {'result_list': result_list})


# enter export to excel

def export_enter_product_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="enter_products.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Enter Products')

    columns = [
        {'name': 'ID', 'width': 2000},
        {'name': 'Product Name', 'width': 6000},
        {'name': 'Quantity', 'width': 2000},
        {'name': 'Created At', 'width': 8000},
    ]

    for col_num, column_data in enumerate(columns):
        ws.write(0, col_num, column_data['name'])
        ws.col(col_num).width = column_data['width']

    enter_products = models.EnterProduct.objects.all()

    for row_num, enter_product in enumerate(enter_products, start=1):
        ws.write(row_num, 0, enter_product.product_id)
        ws.write(row_num, 1, enter_product.product_name)
        ws.write(row_num, 2, enter_product.quantity)
        ws.write(row_num, 3, enter_product.created_at.strftime('%Y-%m-%d %H:%M:%S'))

    wb.save(response)
    return response

# import excel

# excel_file = "malumotlar.xlsx"
# engine = create_engine('sqlite:///malumotlar.db', echo=True)
# Base = declarative_base()

# class Product(Base):
#     __tablename__ = 'products'

#     id = Column(Integer, primary_key=True)
#     quantity = Column(Integer)
#     product_name = Column(String)
#     created_at = Column(DateTime, default=datetime.now)

# Base.metadata.create_all(engine)

# def read_excel_and_save_to_db(file_name):
#     wb = openpyxl.load_workbook(file_name)
#     sheet = wb.active

#     Session = sessionmaker(bind=engine)
#     session = Session()

#     for row in sheet.iter_rows(values_only=True):
#         quantity, product_name, created_at = row
#         product = Product(quantity=quantity, product_name=product_name, created_at=created_at)
#         session.add(product)

#     session.commit()
#     session.close()

# if __name__ == "__main__":
#     read_excel_and_save_to_db(excel_file)

# filterlash

# engine = create_engine('sqlite:///malumotlar.db', echo=True)


# class Product(Base):
#     __tablename__ = 'products'

#     id = Column(Integer, primary_key=True)
#     quantity = Column(Integer)
#     product_name = Column(String)
#     created_at = Column(DateTime, default=datetime.now)


# def filter_products(min_quantity, start_date, end_date):
#     Session = sessionmaker(bind=engine)
#     session = Session()

#     filtered_products = session.query(Product).filter(
#         Product.quantity >= min_quantity,
#         Product.created_at >= start_date,
#         Product.created_at <= end_date
#     ).all()

#     session.close()
#     return filtered_products

# if __name__ == "__main__":
#     min_quantity = 10  
#     start_date = datetime(2023, 1, 1)  
#     end_date = datetime(2023, 12, 31)  

#     filtered_products = filter_products(min_quantity, start_date, end_date)

#     print("Filtered Products:")
#     for product in filtered_products:
#         print(f"ID: {product.id}, Name: {product.product_name}, Quantity: {product.quantity}, Created At: {product.created_at}")

# list enter filterlash


# def list_enter(request):
#     result = search_with_fields(request)
#     try: 
#         enters = models.EnterProduct.objects.filter(**result)

#     except FieldError as err:
#         del result[err.__doc__.split()[3][1:-1]]
#         enters = models.EnterProduct.objects.filter(**result)

#     context = {'enters': pagenator_page(enters, 1, request)}
#     return render(request, 'dashboard/enter/list.html', context)

#required
def list_enter(request):
    name = request.GET.get('name')
    quantity = request.GET.get('quantity')
    created_at = request.GET.get('created_at')
    if name and quantity and created_at:
        enters = models.EnterProduct.objects.filter(
            product__name=name,
            quantity=quantity,
            created_at__gt = created_at,
            created_at__lte = created_at,
        )
    else:
        enters = models.EnterProduct.objects.all()
    context = {'enters':enters}
    return render(request, 'dashboard/enter/list.html', context)

# product filter

# def product_filter(request):
#     result = search_with_fields(request)
#     try: 
#         products = models.Product.objects.filter(**result)

#     except FieldError as err:
#         del result[err.__doc__.split()[3][1:-1]]
#         products = models.Product.objects.filter(**result)

#     context = {'products': pagenator_page(products, 1, request)}
#     return render(request, 'dashboard/products/list.html', context)

def product_filter(request):
    name = request.GET.get('name')
    quantity = request.GET.get('quantity')
    # created_at = request.GET.get('created_at')
    if name and quantity:
        products = models.Product.objects.filter(
            product__name=name,
            quantity=quantity,
            # created_at__gt = created_at,
            # created_at__lte = created_at,
        )
    else:
        products = models.Product.objects.all()
    context = {'products':products}
    return render(request, 'dashboard/products/list.html', context)